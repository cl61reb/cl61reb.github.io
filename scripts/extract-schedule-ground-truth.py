#!/usr/bin/env python3
"""Extracts a per-day owner map from the "Data" tab of the shared custody
schedule workbook, for comparison against the calendar-derived split.

Usage:
    python3 scripts/extract-schedule-ground-truth.py <schedule.xlsx> <range-start> <range-end> <out.json>

The "Data" tab is expected to have columns: Day, Date, Mat, Claire, Comments
(a 1 in Mat or Claire marks who has the kids that date). Output shape:
    { "YYYY-MM-DD": { "owner": "claire"|"parent2"|null, "comment": str|null } }
"""

import sys
import json
import datetime
import openpyxl

def main():
    path, range_start, range_end, out_path = sys.argv[1:5]
    start = datetime.date.fromisoformat(range_start)
    end = datetime.date.fromisoformat(range_end)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Data"]

    out = {}
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 4).value
        if not isinstance(d, datetime.datetime):
            continue
        date = d.date()
        if date < start or date >= end:
            continue
        mat = ws.cell(r, 5).value
        claire = ws.cell(r, 6).value
        comment = ws.cell(r, 7).value
        owner = "parent2" if mat else "claire" if claire else None
        out[date.isoformat()] = {"owner": owner, "comment": comment}

    with open(out_path, "w") as f:
        json.dump(out, f, indent=2, default=str)
    print(f"Wrote {len(out)} days to {out_path}")

if __name__ == "__main__":
    main()
